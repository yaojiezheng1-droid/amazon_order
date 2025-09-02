import sys
import json
import re
from datetime import datetime, timedelta
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image
except ModuleNotFoundError as exc:  # pragma: no cover - dependency missing in tests
    raise SystemExit("openpyxl and pillow are required to run this script") from exc

PRODUCT_START_ROW = 7
COLUMN_MAP = {
    '产品编号': 'A',
    '产品图片': 'B',
    '描述': 'C',
    '数量/个': 'D',
    '单价': 'E',
    '包装方式': 'G',
}


def fill_workbook(template: Path, data: dict):
    """Return workbook filled with ``data`` using ``template``."""
    wb = load_workbook(template)
    ws = wb.active

    for addr, info in data.get('cells', {}).items():
        value = info.get('value', '')
        key = info.get('key', '')
        
        # Handle special date fields
        if key == '日期':
            # Fill with today's date
            value = datetime.now().strftime('%Y年%m月%d日')
        elif key == '交货时间' or key == '交货日期':
            # Extract number from the original value and add to today's date
            original_value = str(value)
            # Look for numbers in the value (could be "15天", "30", "45", etc.)
            numbers = re.findall(r'\d+', original_value)
            if numbers:
                days_to_add = int(numbers[0])
                delivery_date = datetime.now() + timedelta(days=days_to_add)
                value = delivery_date.strftime('%Y年%m月%d日')
            else:
                # If no number found, default to 30 days from today
                delivery_date = datetime.now() + timedelta(days=30)
                value = delivery_date.strftime('%Y年%m月%d日')
        
        ws[addr] = value

    from PIL import Image as PILImage
    row = PRODUCT_START_ROW
    for product in data.get('products', []):
        for key, col in COLUMN_MAP.items():
            if key in product:
                if key == '产品图片':
                    # Use SKU (产品编号) to construct image path, check multiple directories
                    sku = product.get('产品编号', '')
                    # Try products directory first, then accessories directory
                    img_path = template.parent.parent / 'images' / 'products' / f'{sku}.jpg'
                    if not img_path.exists():
                        img_path = template.parent.parent / 'images' / 'accessories' / f'{sku}.jpg'
                    
                    if img_path.exists():
                        try:
                            # Verify image can be opened by Pillow and get size
                            with PILImage.open(img_path) as pil_img:
                                pil_img.verify()
                                orig_width, orig_height = pil_img.size
                            img = Image(str(img_path))
                            # Set row height to 100
                            ws.row_dimensions[row].height = 100
                            # openpyxl row height is in points (1 point = 1/72 inch),
                            # and image.height is in pixels. Excel's default DPI is 96.
                            # 1 point = 1.333 pixels, so 100 points ≈ 133 pixels
                            target_height_px = 133
                            scale = target_height_px / orig_height
                            img.height = target_height_px
                            img.width = int(orig_width * scale)
                            ws.add_image(img, f"{col}{row}")
                        except Exception as e:
                            ws[f"{col}{row}"] = f"[图片错误] {product[key]}: {e}"
                    else:
                        ws[f"{col}{row}"] = f"[图片未找到] {product[key]}"
                else:
                    ws[f"{col}{row}"] = product[key]
        qty = product.get('数量/个')
        price = product.get('单价')
        if qty not in (None, '') and price not in (None, ''):
            ws[f"F{row}"] = f"=D{row}*E{row}"
        row += 1

    footer = data.get('footer', {})
    if 'buyer' in footer:
        ws['B69'] = footer['buyer']
    if 'supplier' in footer:
        ws['E69'] = footer['supplier']
    return wb


def main(argv: list[str]) -> int:
    if len(argv) != 3:
        print("usage: json_PO_excel.py <input.json> <output.xlsx>")
        return 1

    json_path = Path(argv[1])
    out_path = Path(argv[2])
    template = Path(__file__).resolve().parent / 'docs' / 'empty_base_template.xlsx'

    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    wb = fill_workbook(template, data)
    wb.save(out_path)
    return 0


if __name__ == '__main__':
    raise SystemExit(main(sys.argv))
