#!/usr/bin/env python3
"""
Excel to JSON Template Generator with GUI

This script converts Excel files in the format of empty_base_template.xlsx
to JSON template files suitable for the json_template folder.

The script:
1. Reads Excel files with the standard template format
2. Extracts cell data, product information, and footer data
3. Generates individual JSON template files for each product SKU
4. Saves files to the json_template directory with proper formatting

Usage:
    python excel_to_json_template.py                    # Launch GUI (recommended)
    python excel_to_json_template.py input_file.xlsx    # Command line mode
    python excel_to_json_template.py *.xlsx             # Process multiple files

GUI Features:
- File selection with browse dialog
- Folder selection to process all Excel files
- Real-time conversion progress
- Detailed logging and error reporting
- Background processing to keep UI responsive
"""

import json
import re
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl is required. Install it with: pip install openpyxl")
    sys.exit(1)


class ExcelToJsonConverter:
    def __init__(self):
        self.root_dir = Path(__file__).resolve().parent
        self.template_dir = self.root_dir / "json_template"
        self.images_dir = self.root_dir / "images"
        
        # Create template directory if it doesn't exist
        self.template_dir.mkdir(exist_ok=True)
        
        # Load accessory mapping for product names
        self.accessory_map = self._load_accessory_mapping()
        
    def _load_accessory_mapping(self) -> Dict[str, Dict]:
        """Load accessory mapping to get product names"""
        mapping_path = self.root_dir / "docs" / "accessory_mapping.json"
        try:
            with open(mapping_path, "r", encoding="utf-8") as f:
                data = json.load(f)
                return data.get("products", {})
        except FileNotFoundError:
            print(f"Warning: {mapping_path} not found. Product names may not be populated.")
            return {}
    
    def _find_image_path(self, sku: str) -> Optional[str]:
        """Find image path for the given SKU"""
        for sub in ("products", "accessories"):
            img_path = self.images_dir / sub / f"{sku}.jpg"
            if img_path.exists():
                return f"order_generation/images/{sub}/{sku}.jpg"
        return None
    
    def _get_cell_value(self, ws, row: int, col: int) -> str:
        """Get cell value as string"""
        cell = ws.cell(row=row, column=col)
        return str(cell.value or "").strip()
    
    def _extract_cells_data(self, ws) -> Dict[str, Dict[str, str]]:
        """Extract all cell data with keys and values"""
        cells = {}
        
        # Define the cell mapping based on the template structure
        # These are the standard cells that contain metadata
        cell_mappings = {
            # Row 3
            "B3": {"key": "供货商：", "row": 3, "col": 2},
            "G3": {"key": "订单号", "row": 3, "col": 7},
            
            # Row 4
            "B4": {"key": "电话：", "row": 4, "col": 2},
            "G4": {"key": "日期", "row": 4, "col": 7},
            
            # Row 5
            "B5": {"key": "联系人：", "row": 5, "col": 2},
            "G5": {"key": "订单安排人", "row": 5, "col": 7},
            
            # Row 12-17 (various fields)
            "B12": {"key": "进仓地址：", "row": 12, "col": 2},
            "B13": {"key": "付款方式", "row": 13, "col": 2},
            "B14": {"key": "交货时间", "row": 14, "col": 2},
            "F14": {"key": "色卡", "row": 14, "col": 6},
            "G14": {"key": "Logo", "row": 14, "col": 7},
            "B15": {"key": "箱规", "row": 15, "col": 2},
            "F15": {"key": "色卡1", "row": 15, "col": 6},
            "G15": {"key": "Logo1", "row": 15, "col": 7},
            "B16": {"key": "产前确认样", "row": 16, "col": 2},
            "F16": {"key": "色卡2", "row": 16, "col": 6},
            "G16": {"key": "Logo2", "row": 16, "col": 7},
            "B17": {"key": "出货样", "row": 17, "col": 2},
            "F17": {"key": "色卡3", "row": 17, "col": 6},
            "G17": {"key": "Logo3", "row": 17, "col": 7},
            "F18": {"key": "色卡4", "row": 18, "col": 6},
            "G18": {"key": "Logo4", "row": 18, "col": 7},
        }
        
        # Add note fields (A19-A30)
        for i in range(19, 31):
            if i == 19:
                key = "注意事项：1"
            else:
                key = f"{i-18}："
            cell_mappings[f"A{i}"] = {"key": key, "row": i, "col": 1}
        
        # Extract values for mapped cells
        for addr, info in cell_mappings.items():
            value = self._get_cell_value(ws, info["row"], info["col"])
            cells[addr] = {
                "key": info["key"],
                "value": value
            }
        
        return cells
    
    def _find_product_table_start(self, ws) -> int:
        """Find the row where the product table starts"""
        # Look for the header row containing "产品编号", "数量/个", etc.
        for row in range(1, 20):  # Check first 20 rows
            for col in range(1, 8):  # Check columns A-G
                cell_value = self._get_cell_value(ws, row, col)
                if cell_value in ("产品编号", "型号"):
                    return row
        return 7  # Default to row 7 if not found
    
    def _extract_products(self, ws) -> List[Dict[str, Any]]:
        """Extract product data from the worksheet"""
        products = []
        header_row = self._find_product_table_start(ws)
        
        # Define column mapping for product table
        # Based on standard template: A=产品编号, B=产品图片, C=描述, D=数量/个, E=单价, G=包装方式
        col_mapping = {
            1: "产品编号",     # Column A
            2: "产品图片",     # Column B  
            3: "描述",         # Column C
            4: "数量/个",      # Column D
            5: "单价",         # Column E
            7: "包装方式"      # Column G
        }
        
        # Start from the row after header
        row = header_row + 1
        
        while row <= ws.max_row:
            # Get product code from column A
            sku = self._get_cell_value(ws, row, 1)
            
            # Stop if we hit an empty SKU or total row
            if not sku or sku.upper().startswith("TOTAL") or sku == "总计":
                break
            
            product = {}
            
            # Extract data for each column
            for col, field in col_mapping.items():
                value = self._get_cell_value(ws, row, col)
                
                if field == "产品编号":
                    product[field] = value
                elif field == "产品图片":
                    # Try to find image path, use provided value as fallback
                    img_path = self._find_image_path(sku)
                    product[field] = img_path or value
                elif field == "数量/个":
                    # Convert to integer, default to 0
                    try:
                        product[field] = int(float(value)) if value else 0
                    except ValueError:
                        product[field] = 0
                elif field == "单价":
                    # Convert to float
                    try:
                        product[field] = float(value) if value else 0.0
                    except ValueError:
                        product[field] = 0.0
                else:
                    product[field] = value
            
            # Add product name from accessory mapping if available
            if sku in self.accessory_map:
                product["产品名称"] = self.accessory_map[sku]["name"]
            elif "产品名称" not in product:
                product["产品名称"] = ""  # Default empty name
            
            products.append(product)
            row += 1
        
        return products
    
    def _extract_footer(self, ws) -> Dict[str, str]:
        """Extract footer information (buyer, supplier)"""
        footer = {}
        
        # Look for buyer and supplier info around row 69 (standard template)
        try:
            buyer = self._get_cell_value(ws, 69, 2)  # B69
            supplier = self._get_cell_value(ws, 69, 5)  # E69
            
            if buyer:
                footer["buyer"] = buyer
            if supplier:
                footer["supplier"] = supplier
        except:
            pass
        
        return footer
    
    def convert_excel_to_json(self, excel_path: Path) -> List[Path]:
        """Convert Excel file to JSON template(s)"""
        print(f"Processing: {excel_path}")
        
        try:
            wb = load_workbook(excel_path, data_only=True)
            ws = wb.active
            
            # Extract data
            cells = self._extract_cells_data(ws)
            products = self._extract_products(ws)
            footer = self._extract_footer(ws)
            
            if not products:
                print(f"Warning: No products found in {excel_path}")
                return []
            
            generated_files = []
            
            # Group products by SKU to avoid duplicates
            products_by_sku = {}
            for product in products:
                sku = product.get("产品编号")
                if sku:
                    if sku not in products_by_sku:
                        products_by_sku[sku] = product
                    else:
                        # If duplicate SKU, combine quantities
                        existing = products_by_sku[sku]
                        existing["数量/个"] += product.get("数量/个", 0)
            
            # Generate one JSON file per unique product SKU
            for sku, product in products_by_sku.items():
                # Create JSON structure
                json_data = {
                    "cells": cells,
                    "products": [product],
                    "footer": footer
                }
                
                # Save to json_template directory
                output_path = self.template_dir / f"{sku}.json"
                
                with open(output_path, "w", encoding="utf-8") as f:
                    json.dump(json_data, f, ensure_ascii=False, indent=2)
                
                generated_files.append(output_path)
                print(f"  Generated: {output_path}")
            
            return generated_files
            
        except Exception as e:
            print(f"Error processing {excel_path}: {e}")
            return []
    
    def process_files(self, file_patterns: List[str]) -> None:
        """Process multiple Excel files"""
        total_generated = 0
        
        for pattern in file_patterns:
            # Handle both specific files and glob patterns
            if "*" in pattern:
                files = list(Path(".").glob(pattern))
            else:
                files = [Path(pattern)]
            
            for file_path in files:
                if file_path.suffix.lower() in (".xlsx", ".xls"):
                    generated = self.convert_excel_to_json(file_path)
                    total_generated += len(generated)
        
        print(f"\nTotal JSON templates generated: {total_generated}")
        print(f"Output directory: {self.template_dir}")


class ExcelToJsonGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel to JSON Template Converter")
        self.root.geometry("800x600")
        
        # Initialize converter
        self.converter = ExcelToJsonConverter()
        
        # Selected files list
        self.selected_files = []
        
        # Create GUI
        self._create_widgets()
        
    def _create_widgets(self):
        """Create all GUI widgets"""
        # Main frame
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(2, weight=1)
        
        # Title
        title_label = ttk.Label(main_frame, text="Excel to JSON Template Converter", 
                               font=("TkDefaultFont", 16, "bold"))
        title_label.grid(row=0, column=0, pady=(0, 20))
        
        # File selection section
        file_frame = ttk.LabelFrame(main_frame, text="File Selection", padding="10")
        file_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        file_frame.columnconfigure(0, weight=1)
        
        # File selection buttons
        button_frame = ttk.Frame(file_frame)
        button_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        
        ttk.Button(button_frame, text="Select Excel Files", 
                  command=self._select_files).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(button_frame, text="Add Folder", 
                  command=self._select_folder).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(button_frame, text="Clear All", 
                  command=self._clear_files).pack(side=tk.LEFT, padx=(0, 10))
        
        # File list
        list_frame = ttk.Frame(file_frame)
        list_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        list_frame.columnconfigure(0, weight=1)
        list_frame.rowconfigure(0, weight=1)
        
        # Create listbox with scrollbar
        self.file_listbox = tk.Listbox(list_frame, height=6)
        file_scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.file_listbox.yview)
        self.file_listbox.configure(yscrollcommand=file_scrollbar.set)
        
        self.file_listbox.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        file_scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        
        # Control buttons
        control_frame = ttk.Frame(file_frame)
        control_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(10, 0))
        
        ttk.Button(control_frame, text="Remove Selected", 
                  command=self._remove_selected).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(control_frame, text="Convert to JSON", 
                  command=self._convert_files).pack(side=tk.LEFT, padx=(20, 0))
        
        # Progress and output section
        output_frame = ttk.LabelFrame(main_frame, text="Conversion Progress", padding="10")
        output_frame.grid(row=2, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
        output_frame.columnconfigure(0, weight=1)
        output_frame.rowconfigure(1, weight=1)
        
        # Progress bar
        self.progress_var = tk.StringVar(value="Ready to convert files...")
        progress_label = ttk.Label(output_frame, textvariable=self.progress_var)
        progress_label.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 5))
        
        self.progress_bar = ttk.Progressbar(output_frame, mode='determinate')
        self.progress_bar.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # Output text
        self.output_text = scrolledtext.ScrolledText(output_frame, height=15, wrap=tk.WORD)
        self.output_text.grid(row=2, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Status bar
        self.status_var = tk.StringVar(value="Ready")
        status_bar = ttk.Label(main_frame, textvariable=self.status_var, relief=tk.SUNKEN)
        status_bar.grid(row=3, column=0, sticky=(tk.W, tk.E), pady=(10, 0))
        
        # Initial message
        self._log("Excel to JSON Template Converter ready!")
        self._log("Select Excel files to convert to JSON templates.")
        self._log(f"Output directory: {self.converter.template_dir}")
        
    def _log(self, message: str):
        """Log a message to the output text"""
        self.output_text.insert(tk.END, message + "\n")
        self.output_text.see(tk.END)
        self.root.update()
        
    def _select_files(self):
        """Select Excel files to convert"""
        filetypes = [
            ("Excel files", "*.xlsx *.xls"),
            ("All files", "*.*")
        ]
        
        files = filedialog.askopenfilenames(
            title="Select Excel files to convert",
            filetypes=filetypes
        )
        
        if files:
            added_count = 0
            for file_path in files:
                if file_path not in self.selected_files:
                    self.selected_files.append(file_path)
                    self.file_listbox.insert(tk.END, Path(file_path).name)
                    added_count += 1
            
            self._log(f"Added {added_count} file(s)")
            self.status_var.set(f"Selected {len(self.selected_files)} file(s)")
    
    def _select_folder(self):
        """Select all Excel files from a folder"""
        folder = filedialog.askdirectory(title="Select folder containing Excel files")
        
        if folder:
            folder_path = Path(folder)
            excel_files = list(folder_path.glob("*.xlsx")) + list(folder_path.glob("*.xls"))
            
            added_count = 0
            for file_path in excel_files:
                file_str = str(file_path)
                if file_str not in self.selected_files:
                    self.selected_files.append(file_str)
                    self.file_listbox.insert(tk.END, file_path.name)
                    added_count += 1
            
            self._log(f"Added {added_count} file(s) from folder: {folder}")
            self.status_var.set(f"Selected {len(self.selected_files)} file(s)")
    
    def _clear_files(self):
        """Clear all selected files"""
        self.selected_files.clear()
        self.file_listbox.delete(0, tk.END)
        self._log("Cleared all selected files")
        self.status_var.set("Ready")
    
    def _remove_selected(self):
        """Remove selected file from list"""
        selection = self.file_listbox.curselection()
        if selection:
            index = selection[0]
            removed_file = self.selected_files.pop(index)
            self.file_listbox.delete(index)
            self._log(f"Removed: {Path(removed_file).name}")
            self.status_var.set(f"Selected {len(self.selected_files)} file(s)")
    
    def _convert_files(self):
        """Convert selected files to JSON templates"""
        if not self.selected_files:
            messagebox.showwarning("Warning", "Please select Excel files to convert")
            return
        
        # Disable convert button during processing
        for child in self.root.winfo_children():
            self._disable_widgets(child)
        
        # Start conversion in background thread
        thread = threading.Thread(target=self._conversion_worker)
        thread.daemon = True
        thread.start()
    
    def _disable_widgets(self, widget):
        """Recursively disable all widgets"""
        try:
            widget.configure(state='disabled')
        except:
            pass
        for child in widget.winfo_children():
            self._disable_widgets(child)
    
    def _enable_widgets(self, widget):
        """Recursively enable all widgets"""
        try:
            widget.configure(state='normal')
        except:
            pass
        for child in widget.winfo_children():
            self._enable_widgets(child)
    
    def _conversion_worker(self):
        """Background worker for file conversion"""
        try:
            total_files = len(self.selected_files)
            total_generated = 0
            
            self.progress_bar.configure(maximum=total_files)
            
            for i, file_path in enumerate(self.selected_files):
                self.progress_var.set(f"Processing {i+1}/{total_files}: {Path(file_path).name}")
                self.progress_bar.configure(value=i)
                
                self._log(f"\n[{i+1}/{total_files}] Processing: {Path(file_path).name}")
                
                try:
                    generated_files = self.converter.convert_excel_to_json(Path(file_path))
                    total_generated += len(generated_files)
                    
                    if generated_files:
                        self._log(f"  ✓ Generated {len(generated_files)} JSON template(s)")
                        for json_file in generated_files:
                            self._log(f"    - {json_file.name}")
                    else:
                        self._log(f"  ⚠ No templates generated (no products found)")
                        
                except Exception as e:
                    self._log(f"  ✗ Error: {e}")
            
            self.progress_bar.configure(value=total_files)
            self.progress_var.set("Conversion completed!")
            
            # Summary
            self._log(f"\n" + "="*50)
            self._log(f"CONVERSION SUMMARY")
            self._log(f"="*50)
            self._log(f"Files processed: {total_files}")
            self._log(f"JSON templates generated: {total_generated}")
            self._log(f"Output directory: {self.converter.template_dir}")
            self._log(f"="*50)
            
            if total_generated > 0:
                messagebox.showinfo("Success", 
                    f"Conversion completed!\n\n"
                    f"Files processed: {total_files}\n"
                    f"JSON templates generated: {total_generated}\n"
                    f"Output directory: {self.converter.template_dir}")
            else:
                messagebox.showwarning("Warning", 
                    f"Conversion completed but no JSON templates were generated.\n"
                    f"Please check the Excel files contain valid product data.")
            
            self.status_var.set(f"Completed: {total_generated} templates generated")
            
        except Exception as e:
            self._log(f"\nUnexpected error: {e}")
            messagebox.showerror("Error", f"Conversion failed: {e}")
            self.status_var.set("Error occurred")
        
        finally:
            # Re-enable widgets
            for child in self.root.winfo_children():
                self._enable_widgets(child)


def main():
    # Check if GUI mode should be used
    if len(sys.argv) == 1:
        # No command line arguments - launch GUI
        try:
            root = tk.Tk()
            app = ExcelToJsonGUI(root)
            root.mainloop()
        except Exception as e:
            print(f"GUI Error: {e}")
            print("Falling back to command line mode...")
            print("Usage: python excel_to_json_template.py <excel_file1> [excel_file2] ...")
            sys.exit(1)
    else:
        # Command line arguments provided - use CLI mode
        print("Usage: python excel_to_json_template.py <excel_file1> [excel_file2] ...")
        print("       python excel_to_json_template.py *.xlsx")
        print("       python excel_to_json_template.py  # Launch GUI")
        
        converter = ExcelToJsonConverter()
        converter.process_files(sys.argv[1:])


if __name__ == "__main__":
    main()
