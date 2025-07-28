import json
from typing import Dict, Any, List
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage


START_ROW = 7  # first product row in the template
PLACEHOLDER_ROWS = 3  # number of placeholder product rows in the template


def fill_cells(ws, cells: Dict[str, Any]) -> None:
    """Write values from the json ``cells`` mapping into the worksheet.

    ``cells`` now maps cell addresses to objects with ``value`` and
    optionally ``key``.  Only the ``value`` field is written into the
    worksheet, but keeping the ``key`` in the JSON makes it easier to
    understand where the value came from when converting back from Excel.
    """
    for address, meta in cells.items():
        if isinstance(meta, dict):
            value = meta.get("value", "")
        else:
            value = meta
        ws[address] = value


def insert_products(ws, products: List[Dict[str, Any]]) -> None:
    """Insert product rows into the worksheet starting at START_ROW."""
    ws.delete_rows(START_ROW, PLACEHOLDER_ROWS)
    ws.insert_rows(START_ROW, len(products))

    for idx, item in enumerate(products):
        r = START_ROW + idx
        ws.cell(row=r, column=1, value=item.get("产品编号", ""))
        img_path = item.get("产品图片")
        if img_path:
            try:
                ws.add_image(XLImage(img_path), f"B{r}")
            except Exception:
                ws.cell(row=r, column=2, value=img_path)
        ws.cell(row=r, column=3, value=item.get("描述", ""))
        ws.cell(row=r, column=4, value=item.get("数量/个", ""))
        ws.cell(row=r, column=5, value=item.get("单价", ""))
        ws.cell(row=r, column=6, value=f"=E{r}*D{r}")
        ws.cell(row=r, column=7, value=item.get("包装方式", ""))

    total_row = START_ROW + len(products)
    ws.cell(row=total_row, column=4, value=f"=SUM(D{START_ROW}:D{total_row-1})")
    ws.cell(row=total_row, column=6, value=f"=SUM(F{START_ROW}:F{total_row-1})")


def create_order_workbook(data: Dict[str, Any], template: str, output: str) -> None:
    wb = load_workbook(template)
    ws = wb.active

    fill_cells(ws, data.get("cells", {}))
    insert_products(ws, data.get("products", []))

    footer = data.get("footer", {})
    if footer:
        if "buyer" in footer:
            ws["B69"] = footer["buyer"]
        if "supplier" in footer:
            ws["E69"] = footer["supplier"]

    wb.save(output)


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Fill order template spreadsheet")
    parser.add_argument("json", help="Path to order data json")
    parser.add_argument("output", help="Output Excel file path")
    parser.add_argument("--template", default="docs/template_order_excel_1.xlsx", help="Template workbook path")
    args = parser.parse_args()

    with open(args.json, "r", encoding="utf-8") as f:
        data = json.load(f)

    create_order_workbook(data, args.template, args.output)
