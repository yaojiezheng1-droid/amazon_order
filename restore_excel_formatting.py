import openpyxl
import os
import shutil
from pathlib import Path

def restore_excel_formatting(original_file, modified_file, output_file):
    """
    Restore Excel formatting by copying data from modified file to original formatted file
    """
    try:
        # Load the original file (with formatting)
        original_wb = openpyxl.load_workbook(original_file)
        
        # Load the modified file (with updated data but no formatting)
        modified_wb = openpyxl.load_workbook(modified_file, data_only=True)
        
        # Process each worksheet
        for sheet_name in modified_wb.sheetnames:
            if sheet_name in original_wb.sheetnames:
                original_ws = original_wb[sheet_name]
                modified_ws = modified_wb[sheet_name]
                
                # Copy values from modified sheet to original sheet (preserving formatting)
                for row in modified_ws.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            original_cell = original_ws.cell(row=cell.row, column=cell.column)
                            original_cell.value = cell.value
        
        # Save the result
        original_wb.save(output_file)
        print(f"Successfully restored formatting for {os.path.basename(output_file)}")
        return True
        
    except Exception as e:
        print(f"Error processing {os.path.basename(modified_file)}: {str(e)}")
        return False

def main():
    # Define paths
    original_dir = r"c:\Users\Cheng\Desktop\amazon_order\order_generation\PO_excel"
    
    # Check if we should use the restored folder as source (if it exists and has content)
    restored_dir = r"c:\Users\Cheng\Desktop\amazon_order\order_generation\PO_excel_restored"
    selected_dir = r"c:\Users\Cheng\Desktop\amazon_order\order_generation\PO_excel_selected"
    
    # Use restored folder if it exists and has files, otherwise use selected folder
    if os.path.exists(restored_dir) and os.listdir(restored_dir):
        modified_dir = restored_dir
        print("Using previously restored files as source...")
    else:
        modified_dir = selected_dir
        print("Using selected files as source...")
    
    output_dir = r"c:\Users\Cheng\Desktop\amazon_order\order_generation\PO_excel_restored"
    
    # Create output directory
    os.makedirs(output_dir, exist_ok=True)
    
    # Files that need formatting restoration (small file sizes indicating lost formatting)
    problem_files = [
        "17-A1KN-KJGW.xlsx",
        "2EC-1-1.xlsx",
        "2EC-Blue.xlsx", 
        "2EC-Green.xlsx",
        "2EC-Pink.xlsx",
        "2EC-Yellow.xlsx",
        "48-82P3-QSFG.xlsx",
        "7S-HA5T-5D0X.xlsx",
        "AMCB-01.xlsx",
        "AMCB-black.xlsx",
        "AMCB-Blue.xlsx",
        "AMCB-Pink.xlsx",
        "B10-GTS2-4JZ-BK-3.xlsx",
        "EC404.xlsx"
    ]
    
    print("Starting Excel formatting restoration...")
    print("=" * 50)
    
    successful = 0
    failed = 0
    
    for filename in problem_files:
        original_file = os.path.join(original_dir, filename)
        modified_file = os.path.join(modified_dir, filename)
        output_file = os.path.join(output_dir, filename)
        
        if os.path.exists(original_file) and os.path.exists(modified_file):
            if restore_excel_formatting(original_file, modified_file, output_file):
                successful += 1
            else:
                failed += 1
        else:
            print(f"Missing files for {filename}")
            failed += 1
    
    # Copy files that didn't need restoration
    print("\nCopying files that retained formatting...")
    all_files = os.listdir(modified_dir)
    for filename in all_files:
        if filename.endswith('.xlsx') and filename not in problem_files:
            src = os.path.join(modified_dir, filename)
            dst = os.path.join(output_dir, filename)
            shutil.copy2(src, dst)
            print(f"Copied {filename}")
    
    print("=" * 50)
    print(f"Restoration complete!")
    print(f"Successfully restored: {successful} files")
    print(f"Failed: {failed} files")
    print(f"Output directory: {output_dir}")

if __name__ == "__main__":
    main()
