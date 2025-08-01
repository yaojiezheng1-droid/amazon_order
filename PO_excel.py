import json
from typing import Dict, Any, List
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import column_index_from_string, get_column_letter

START_ROW = 7  # first product row in the template
PLACEHOLDER_ROWS = 3  # number of placeholder product rows in the template
BUYER_ROW = 69  # base row for buyer name (before any inserted rows)

def validate_cells(ws, cells: Dict[str, Any]) -> None:
    """Ensure every address exists and its label matches the JSON ``key``.

    The JSON ``cells`` mapping stores entries like ``{"B3": {"key": "供货商：", "value": "..."}}``.
    The value for ``key`` is expected to match the text in either the target cell
    or the cell to its immediate left.  A ``ValueError`` is raised if a mismatch
    is found so the template can be updated before continuing.
    """
    for address, meta in cells.items():
        # Make sure the address is valid
        try:
            cell = ws[address]
        except ValueError as exc:
            raise ValueError(f"Invalid cell address '{address}' in JSON") from exc

        if not isinstance(meta, dict):
            continue
        key = meta.get("key")
        if not key:
            continue

        # First check the cell itself
        if cell.value == key:
            continue

        # Then check the cell to the left if possible
        col_letters = ''.join(filter(str.isalpha, address))
        row_numbers = ''.join(filter(str.isdigit, address))
        col_idx = column_index_from_string(col_letters)
        if col_idx > 1:
            left_addr = f"{get_column_letter(col_idx - 1)}{row_numbers}"
            if ws[left_addr].value == key:
                continue
        # Neither cell matched; raise an error
        raise ValueError(
            f"Key '{key}' for cell '{address}' does not match template (found '{cell.value}')."
        )

def fill_cells(ws, cells: Dict[str, Any]) -> None:
    """Write values from the JSON ``cells`` mapping into the worksheet."""
    validate_cells(ws, cells)
    for address, meta in cells.items():
        value = meta.get("value") if isinstance(meta, dict) else meta
        ws[address] = value

def insert_products(ws, products: List[Dict[str, Any]]) -> int:
    """Insert product rows into the worksheet starting at ``START_ROW``.

    Returns the number of rows inserted which is used to adjust footer
    positions below the product table.
    """
    extra = max(0, len(products) - PLACEHOLDER_ROWS)
    if extra:
        ws.insert_rows(START_ROW + PLACEHOLDER_ROWS, extra)

    row_count = max(len(products), PLACEHOLDER_ROWS)
    for idx in range(row_count):
        r = START_ROW + idx
        item = products[idx] if idx < len(products) else {}
        ws.cell(row=r, column=1, value=item.get("产品编号", ""))
        img_path = item.get("产品图片")
        if img_path:
            try:
                ws.add_image(XLImage(img_path), f"B{r}")
            except Exception:
                # If the image can't be loaded, fall back to writing the path
                ws.cell(row=r, column=2, value=img_path)
        else:
            ws.cell(row=r, column=2, value="")
        name = item.get("产品名称", "").strip()
        desc = item.get("描述", "").strip()
        if name:
            desc = f"{name} {desc}" if desc else name
        ws.cell(row=r, column=3, value=desc)
        ws.cell(row=r, column=4, value=item.get("数量/个", ""))
        ws.cell(row=r, column=5, value=item.get("单价", ""))
        ws.cell(row=r, column=6, value=f"=E{r}*D{r}")
        ws.cell(row=r, column=7, value=item.get("包装方式", ""))

    total_row = START_ROW + row_count
    ws.cell(row=total_row, column=4, value=f"=SUM(D{START_ROW}:D{total_row-1})")
    ws.cell(row=total_row, column=6, value=f"=SUM(F{START_ROW}:F{total_row-1})")
    return extra

def create_order_workbook(data: Dict[str, Any], template: str, output: str) -> None:
    wb = load_workbook(template)
    ws = wb.active

    fill_cells(ws, data.get("cells", {}))
    offset = insert_products(ws, data.get("products", []))

    footer = data.get("footer", {})
    if footer:
        buyer_cell = f"B{BUYER_ROW + offset}"
        supplier_cell = f"E{BUYER_ROW + offset}"
        if "buyer" in footer:
            ws[buyer_cell] = footer["buyer"]
        if "supplier" in footer:
            ws[supplier_cell] = footer["supplier"]

    wb.save(output)

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Fill purchase order template spreadsheet")
    parser.add_argument("json", help="Path to order data json")
    parser.add_argument("output", help="Output Excel file path")
    parser.add_argument(
        "--template",
        default="order_generation/docs/empty_base_template.xlsx",
        help="Template workbook path",
    )
    args = parser.parse_args()

    with open(args.json, "r", encoding="utf-8") as f:
        data = json.load(f)

    create_order_workbook(data, args.template, args.output)
