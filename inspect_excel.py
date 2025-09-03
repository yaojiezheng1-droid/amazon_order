#!/usr/bin/env python3
import openpyxl

# Load the Excel file
wb = openpyxl.load_workbook('order_generation/docs/PO_import_empty.xlsx')
ws = wb.active

print("Headers from row 1:")
headers = []
for col in range(1, ws.max_column + 1):
    cell_value = ws.cell(1, col).value
    headers.append(cell_value)
    print(f"Col {col}: {cell_value}")

print("\nSample data from row 2:")
for col in range(1, min(15, ws.max_column + 1)):
    cell_value = ws.cell(2, col).value
    print(f"Col {col}: {cell_value}")

print(f"\nTotal columns: {ws.max_column}")
print(f"Total rows: {ws.max_row}")

print("\nAll headers from row 2:")
for col in range(1, ws.max_column + 1):
    cell_value = ws.cell(2, col).value
    print(f"{col}: {cell_value}")
